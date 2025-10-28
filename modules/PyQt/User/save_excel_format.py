from modules.user.class_utils import Class_Utils

from openpyxl import Workbook, drawing
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Color, PatternFill, colors
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
import pathlib

# Imorting the necessary modules
try:
		from openpyxl.cell import get_column_letter
except ImportError:
		from openpyxl.utils import get_column_letter
		from openpyxl.utils import column_index_from_string

import pandas as pd
import requests, copy
from PIL import Image

from info import Info_SW as INFO
import traceback
from modules.logging_config import get_plugin_logger



# 인자 없이 호출하면 자동으로 현재 모듈 이름(파일 이름)을 사용
logger = get_plugin_logger()

class Save_Excel_format_작지(Class_Utils):
	def __init__(self, dataObj:dict, **kwargs):
		self.dataObj = dataObj
		self.init_attributes(**kwargs)
		self.image_width = 600
		self.image_height = 400

		self.dataTailList = [ 
			'고객요청사항',
			'고객성향',
			'특이사항',
			'집중점검항목',
			'검사요청사항',	
		]
		self.skip_align= [8]

		self.의장도_dict = {

		}


		self.base_강조_Font = Font(size=20, italic = False, bold = True)
		self.fontStyle = {
			"A1" : Font(size=20, italic = False, bold = True),
			"A3" :self.base_강조_Font,
		}

		self.base_rowHeight = 24
		self.base_colWidth = 12

		self.gray_fill = PatternFill(
				fill_type='solid',
				fgColor= colors.Color(rgb='00E0E0E0')
			   )
		self.cell_fills_dict = {
			'A1' : self.gray_fill,
			'A3' : self.gray_fill,
			'A9' :  self.gray_fill,
			'B9' :  self.gray_fill,
			'C9' :  self.gray_fill,
			'E9' :  self.gray_fill,
			'F9' :  self.gray_fill,
			'G9' :  self.gray_fill,
			'J9' :  self.gray_fill,		
		}

		self.cell_상세Process_fills = [
			PatternFill(
				fill_type='solid',
				fgColor= colors.Color(rgb='00FFE5CC')
			   ),
			PatternFill(
				fill_type='solid',
				fgColor= colors.Color(rgb='00FFCCCC')
			   ),
			PatternFill(
				fill_type='solid',
				fgColor= colors.Color(rgb='00CCCCFF')
			   ),
			PatternFill(
				fill_type='solid',
				fgColor= colors.Color(rgb='00CCE5FF')
			   ),
			PatternFill(
				fill_type='solid',
				fgColor= colors.Color(rgb='00FFCCFF')
			   ),
			PatternFill(
				fill_type='solid',
				fgColor= colors.Color(rgb='00CCFFCC')
			   ),
			# PatternFill(
			# 	fill_type='solid',
			# 	fgColor= colors.Color(rgb='00FFFF90')
			#    ),
			# PatternFill(
			# 	fill_type='solid',
			# 	fgColor= colors.Color(rgb='00FFFF90')
			#    ),
		]

	def save_to_excel_from_dict(self) -> str:
		wb = Workbook()

		if hasattr(self, '의장도_fk_datas'):
			self.의장도_fk_datas:dict
			if self.의장도_fk_datas:

				for key, obj in self.의장도_fk_datas.items():
					if isinstance(obj, dict):
						self.의장도_dict[obj.get('title')] = obj.get('file')
			self.write_의장도_sheet(wb , self.의장도_dict)


		self.data = {
			"A1" : "작 업 지 침 서",
			"A3" : self.dataObj.get('제목'),
			"A4" : 'Prjo No',
			'B4' : self.dataObj.get('Proj_No'),
			"C4" : '고객사',
			"D4" : self.dataObj.get('고객사'),
			'E4' : '구분',
			'F4' : self.dataObj.get('구분'),
			'A5' : 'EL수량',
			'B5' : self.dataObj.get('수량'),
			'C5' : '납기일(최초)',
			'D5' : self.dataObj.get('납기일'),
			'E5' : '담당',
			'F5' : self.dataObj.get('담당'),

			'A6' : '영업담당',
			'B6' : self.dataObj.get('영업담당'),
			'C6' : '작성자',
			'D6' : self.dataObj.get('작성자'),
			'E6' : '작성일',
			'F6' : self.dataObj.get('작성일'),

			'A7' : 'EL수량',
			'B7' : self.el_info_dict.get('수량', '') if hasattr(self, 'el_info_dict') else '',
			'C7' : '운행층수',
			'D7' : self.el_info_dict.get('운행층수','') if hasattr(self, 'el_info_dict') else '',
			'E7' : '건물주소',
			'F7' : self.el_info_dict.get('건물주소','') if hasattr(self, 'el_info_dict') else '',

			'A8' : '  1.작 업 내 용',

		}

		self.작업내용_header = {
			'A9' : '순서',
			'B9' : '적용부품',
			'C9' : '적용패널',
			'E9' : 'Material',
			'F9' : '대표Process',
			'G9' : '상세Process',
			'J9' : '비고',
		}

		
		self.col_width_dict = {
			'A' : int(self.base_colWidth*0.75),
			'B' : int(self.base_colWidth),
			'C' : int(self.base_colWidth),
			'D' : int(self.base_colWidth),
			'E' : int(self.base_colWidth),
			'F' : int(self.base_colWidth)*2,
			'G' : int(self.base_colWidth),
			'H' : int(self.base_colWidth),
			'I' : int(self.base_colWidth),
			'J' : int(self.base_colWidth),
			'K' : int(self.base_colWidth),			
		}

		ws = wb.active

		self.merge(ws)
		self.write_cell(ws, self.data)
		
		self._set_col_span(ws)

		self._set_all_alignment(ws)
		self._set_font_style(ws)
		self._set_border_all(ws)
		self._set_column_width(ws, self.col_width_dict)
		self._set_row_width(ws)
		self._set_cell_fill(ws, self.cell_fills_dict)


		try:
			제목 = self.dataObj.get('제목','')
			Proj_No = self.dataObj.get('Proj_No','')
			fName = str(pathlib.Path.home()) +f"/{제목}({Proj_No}).xlsx"

			ws.title =  제목 + f"({Proj_No})"
			wb.save(filename= fName)
			wb.close()
			return fName
		except Exception as e:

			wb.close()
			return ''

	def write_의장도_sheet(self, wb:Workbook, 의장도_dict:dict) -> None:
		def title_cell_style_적용(cell, title_Fill, title_Font):
			cell.fill = title_Fill
			cell.font = title_Font

		ws = wb.create_sheet('의장도', index=1)
		ws:Worksheet
		ws.column_dimensions['A'].width = self.image_width
		ws.column_dimensions['B'].width = self.image_width

		title_Fill = PatternFill(
				fill_type='solid',
				fgColor= colors.Color(rgb='00FFFF66')
			   )
		title_Font = self.base_강조_Font

		positionList = [ ('A1','A2'), ('B1','B2'), ('A4','A5'), ('B4','B5'),
				  		('A7','A8'), ('B7','B8'), ('A10','A11'), ('B10','B11')]
		for idx, (name, url) in enumerate(의장도_dict.items()):
			if url is None : continue
			position = positionList[idx][0]
			ws[position].value = name
			ws.row_dimensions[int(position[1:])].height = self.base_rowHeight*1.5
			
			title_cell_style_적용( ws[position], title_Fill, title_Font )
			self.insert_img_from_URL(ws, url, positionList[idx][1] )


	def merge(self, ws:Worksheet) -> None:
		""" cell merge by format"""
		ws.merge_cells("A1:F2")
		ws.merge_cells("A3:F3")
		ws.merge_cells("A8:K8")
		
		if hasattr(self, 'process'):
			### table header merge
			ws.merge_cells("C9:D9")
			ws.merge_cells("G9:I9")
			ws.merge_cells("J9:K9")
			for row수, _ in enumerate(self.process):
				rowNo = 9+row수+1
				ws.merge_cells(f"C{rowNo}:D{rowNo}")
				ws.merge_cells(f"G{rowNo}:I{rowNo}")
				ws.merge_cells(f"J{rowNo}:K{rowNo}")
		
		last_row = 8 + 1 + len(self.process)

		ws.merge_cells(f"A{last_row+1}:K{last_row+1}")
		ws[f"A{last_row+1}"].value = ' 2.참 고 사 항 (렌더링)'
		self.skip_align.append(last_row+1)
		self.image_row = last_row+2
		ws.merge_cells(f"A{self.image_row}:K{self.image_row}")

		for idx, name in enumerate(self.dataTailList):
			ws.merge_cells( f"A{self.image_row+1+idx}:D{self.image_row+1+idx}")
			ws.merge_cells( f"E{self.image_row+1+idx}:K{self.image_row+1+idx}")
			ws[f"A{self.image_row+1+idx}"].value = name

	def write_cell(self, ws:Worksheet, data:dict) -> None:
		for cellName, value in data.items():
			ws[cellName].value = value

		for cellName, value in self.작업내용_header.items():
			ws[cellName].value = value

		if hasattr(self, 'process'):
			for row, obj in enumerate(self.process):
				for cellName, valueName in self.작업내용_header.items():
					cell = ws[f"{cellName[0]}{int(cellName[1:]) + row+1}"]
					if valueName == '순서' : cell.value = row+1
					else: 
						match cellName:
							case 'F9'|'G9':
								cell.value = self._getValue_의장 ( obj.get(valueName, '') )
							case _:
								cell.value =  obj.get(valueName, '')
		
		if 'Rendering_URL' in self.dataObj and ( url:= self.dataObj['Rendering_URL']):

			self.insert_img_from_URL( ws, url, f"A{self.image_row}" )
		
		for cellName, value in self._get_data_tail().items():
			ws[cellName].value = value
		
	def insert_img_from_URL(self, ws:Worksheet, url:str, position:str) -> None:
		if url is None or not url : return 
		if 'http' not in url : url = INFO.URI+url
		image_open =  Image.open ( requests.get( url, stream=True).raw )
		img = drawing.image.Image(image_open)
		img.width = self.image_width 
		img.height = self.image_height
		ws.add_image(img, position)		
		ws.row_dimensions[int(position[1:])].height = img.height
	
	def _get_data_tail(self) -> dict:
		row = self.image_row +1
		obj = {}
		for idx, name in enumerate(self.dataTailList):
			obj[f"E{row+idx}"] = self.dataObj.get(name)
		return obj
	
	# https://stackoverflow.com/questions/26671581/horizontal-text-alignment-in-openpyxl
	def _set_all_alignment(self, ws:Worksheet, h_align:str='center', v_align:str='center') ->None:
		tot_rows = ws.max_row+1 #get max row number
		tot_cols = ws.max_column #get max column number

		cols = range(1,tot_cols) #turns previous variables into iterables
		rows = range(1,tot_rows) 

		for c in cols:
			for r in rows:
				if r in self.skip_align : continue
				ws.cell(row=r, column=c).alignment = Alignment(horizontal= h_align, vertical= v_align)

	def _set_font_style(self, ws:Worksheet):
		for cellName, fontValue in self.fontStyle.items():
			ws[cellName].font = fontValue

	def _set_border_all(self, ws:Worksheet):		
		""" all 적용"""
		tot_rows = ws.max_row+1 #get max row number
		tot_cols = ws.max_column #get max column number

		cols = range(1,tot_cols) #turns previous variables into iterables
		rows = range(1,tot_rows) 
		border_format = Side(border_style='thin', color='111111')
		for c in cols:
			for r in rows:
				if r in self.skip_align : continue
				ws.cell(row=r, column=c).border = Border(top=border_format, left=border_format,right=border_format,bottom=border_format)

	# https://stackoverflow.com/questions/60248319/how-to-set-column-width-to-bestfit-in-openpyxl
	def _set_column_width(self, ws:Worksheet, COL_WIDTH_DICT:dict={}):
		""" if COL_WIDTH_DICT 가 있으면 FIXED HEIGHT 적용 , 없으면 AUTO 적용"""
		if COL_WIDTH_DICT:
			for col_letter, width in COL_WIDTH_DICT.items():
				ws.column_dimensions[col_letter].width = width
			return 

		for column_cells in ws.columns:
			new_column_length = max(len(str(cell.value)) for cell in column_cells)
			new_column_letter = (get_column_letter(column_cells[0].column))
			if new_column_length > 0:
				ws.column_dimensions[new_column_letter].width = new_column_length*1.23


	def _set_row_width(self, ws:Worksheet):
		for rowNo, row in enumerate(ws.iter_rows()):
			if rowNo > 9 and rowNo < self.image_row-1 :
				for colNo, cell in enumerate(row):
					if colNo == 7 :
						value = ws.cell(row=rowNo, column=colNo).value						
						ws.row_dimensions[rowNo].height = self._get_RowHeight_by_value(value)

	def _set_cell_fill(self, ws:Worksheet, FILL_DICT:dict):
		for rowNo, row in enumerate(ws.iter_rows()):
			if rowNo > 9 and rowNo < self.image_row-1 :
				for colNo, cell in enumerate(row):
					if colNo == 7 :
						value = ws.cell(row=rowNo, column=colNo).value		
						index = self._get_column_중요도(ws, data=value)
						if index >= 0 and index <= len(self.cell_상세Process_fills)-1 :
							ws.cell(row=rowNo, column=colNo).fill= self.cell_상세Process_fills[index]
		if FILL_DICT:
			for cellName, fillValue in FILL_DICT.items():
				ws[cellName].fill = fillValue


	def _set_col_span(self, ws:Worksheet, colNos :list= [2,3,5,6,7,10]):
		for colNo, col in enumerate(ws.iter_cols() ):
			startNo = 0
			spanCount = 0
			preValue = ''
			if colNo in colNos :
				for rowNo, row in enumerate(ws.iter_rows()):
					if rowNo > 9 and rowNo < self.image_row-1 :
						original_value =  ws.cell(row=rowNo, column=colNo).value 
						cur_value =  self._getValue_정리( original_value )

						if rowNo == 10 : 
							startNo = rowNo
							preValue = cur_value
						else:
							if preValue == cur_value :
								spanCount +=1
							else:
								if spanCount:

									if colNo == 7 or colNo == 9:
										for cnt in range(spanCount):
											ws.unmerge_cells(start_row= startNo+cnt, start_column= colNo, 
														end_row=startNo+cnt, end_column=self._get_End_column(colNo) )
									
									ws.merge_cells( start_row= startNo, start_column= colNo, 
												end_row=startNo+spanCount, end_column=self._get_End_column(colNo))
									# ws.cell(row=startNo,column=colNo).value = original_value
								preValue = cur_value
								spanCount = 0
								startNo = rowNo

				if spanCount:
					if colNo == 7:
						for cnt in range(spanCount):
							ws.unmerge_cells(start_row= startNo, start_column= colNo, 
										end_row=startNo+cnt, end_column=self._get_End_column(colNo) )
					ws.merge_cells( start_row= startNo, start_column= colNo, 
									end_row=startNo+spanCount, end_column=self._get_End_column(colNo))
					# ws.cell(row=startNo,column=colNo).value = original_value

	def _getValue_정리(self, value:str ) -> str:
		""" value에 대해 space, /n을 제거한 순수한 string return"""
		if not value : return ''
		new_value = value.replace(' ', '')
		return new_value.replace('\n', '')

	def _get_End_column(self, colNo:int) -> int:
		match colNo:
			case 3: return 4
			case 7: return 9
			case 10: return 11
			case _: return colNo

	def _getValue_의장 (self, value:str) -> str:
		""" 의장명을 받아서 +(구분자) 뒤에 \n 추가하여 return"""
		if not value : return ''
		newValue = value.replace('\n', '')
		return newValue.replace('+', '+\n')

	def _get_RowHeight_by_value(self, value:str) -> int:
		""" value를 받아 +(구분자) 수만큼 height 변경"""
		if not value : return self.base_rowHeight
		cnt = value.count('+')
		return self.base_rowHeight * (cnt+1)


	def _get_column_중요도(self, col_key:str='상세Process', data:str='' ) -> int|None:
		if not data : return -1
		col_datas_list=[]
		for idx, obj in enumerate(self.process):
			obj:dict
			col_datas_list.append(  self._getValue_정리 (obj.get('상세Process') )  )

		col_datas_unique = list(set(col_datas_list))
		self.col_len_sorted = sorted( col_datas_unique, key=len,  reverse=True)


		return self.col_len_sorted.index( self._getValue_정리(data))

	# def _get_last_RowNo(self, ws, col_letter:str) -> int :
	# 	if col_letter :
	# 		return len(ws[col_letter])
	# 		for cell in ws[col_letter]:

	# 			if cell.value : 


	# 		return len([cell for cell in ws[col_letter] if cell.value is not None])

		# https://stackoverflow.com/questions/33541692/how-to-find-the-last-row-in-a-column-using-openpyxl-normal-workbook
		# for col in range(1, ws.max_column + 1):
		# 	col_letter = get_column_letter(col)
		# 	max_col_row = len([cell for cell in ws[col_letter] if cell.value is not None])


		
