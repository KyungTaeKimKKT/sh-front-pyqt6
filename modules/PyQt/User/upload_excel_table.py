from modules.user.class_utils import Class_Utils

from openpyxl import Workbook, drawing, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl_image_loader import SheetImageLoader
import pandas as pd
from datetime import datetime

import modules.user.utils as Utils
from info import Info_SW as INFO
import traceback
from modules.logging_config import get_plugin_logger



# 인자 없이 호출하면 자동으로 현재 모듈 이름(파일 이름)을 사용
logger = get_plugin_logger()

class Upload_Excel_작지(Class_Utils):
	def __init__(self, fName:str, **kwargs):
		self.init_attributes(**kwargs)
		self.fName = fName 
		self.process_table_header=  {
			2: '적용부품',
			3: '적용패널',
			5: 'Material',
			6: 'Process',
			10: '비고',
		}
		self.dataObj = {'id': -1}
		self.process_result = []
		self.checkList =  ['고객요청사항','고객성향','특이사항','집중점검항목','검사요청사항']

		self.read(fName)
		self.get_table_process_datas(self.ws)
		self.get_datas(self.ws)
		self.dataObj['process_fks']= self.process_result
		self.get_image(self.ws)

	def _getDataObj(self):
		return self.dataObj

	def read(self, fName):
		self.wb = load_workbook(fName)
		self.ws = self.wb.worksheets[0]

	def get_table_process_datas(self, ws:Worksheet):
		for rowNo in range(1,ws.max_row):
			value:str = ws.cell(row=rowNo, column=1).value
			if value and isinstance(value, str):
				if '작업내용' in value.replace(' ',''):
					table_start_row = rowNo+2
				if '참고사항' in value.replace(' ',''):
					table_end_row = rowNo-1
					self.tail_start = rowNo+2

		표시순서 = 0
		for rowNo in range(1,ws.max_row):
			process = {'id' : -1}
			if rowNo >= table_start_row and rowNo <= table_end_row:
				for colNo, head_Name in self.process_table_header.items():

					value:str = ws.cell(row=rowNo, column=colNo).value
					if value is None and '적용패널' not in head_Name :
						rowIndex = rowNo
						while not value:
							rowIndex -= 1
							if rowIndex < table_start_row : break
							value = ws.cell(row=rowIndex, column=colNo).value
					match self.process_table_header[colNo]:
						case 'Process':					
							valueList = value.split(';')
							if len(valueList) == 2:
								process['대표Process'] = value.split(';')[0]
								process['상세Process'] = value.split(';')[1]
							else:
								process['대표Process'] = value.split(';')[0]
								process['상세Process'] = value.split(';')[0]
						case 'Material':
							if value.upper() not in INFO.Material_Widget_itmes:
								process['Material'] = '해당없음' if value == '-' else '기타'
							else :
								process['Material'] = value.upper()
						case _:
							process[self.process_table_header[colNo]] = value
				process['표시순서'] = 표시순서
				표시순서 += 1
				if 'front' in process['적용부품'].lower():
					process['적용부품'] = 'Front'
					process['적용패널'] = 'Front Panel' if not process['적용패널'] else process['적용패널']
				elif 'car door' in process['적용부품'].lower():
					# process['적용부품'] = 'Front'
					process['적용패널'] = 'Car Door' if not process['적용패널'] else process['적용패널']
				self.process_result.append(process)

	def get_datas(self, ws:Worksheet):
		title:str = ws['A3'].value
		titleList = title.split('\n')
		self.dataObj['제목'] = titleList[0]
		if len(titleList) >= 2:
			sub_title =  title.split('\n')[1]
			sub_title = sub_title.replace('(','').replace(')','')
			if '수량' in sub_title and '대' in sub_title:
				if ( 수량:= Utils.get_int_from_string( sub_title[sub_title.index('수량'):sub_title.index('대')] )):
					self.dataObj['수량'] = 수량
			if '납기' in sub_title :
				납기일str = sub_title[sub_title.index('납기'):].split(':')[1].strip()
				납기일_obj = datetime.strptime(납기일str, '%y/%m/%d')
				self.dataObj['납기일'] = 납기일_obj
		try:
			self.dataObj['작성일'] = ws['B6'].value
			if ( 고객사 := Utils.check_str_contains_listElements(ws['E6'].value.replace(' ',''), INFO.고객사_Widget_items )):
				self.dataObj['고객사'] = 고객사
			else :
				self.dataObj['고객사'] = ws['E6'].value
		except:
			pass
		self.dataObj['작성자'] = value if (value:=ws['B7'].value ) else ''
		self.dataObj['담당'] = value if (value:=ws['E7'].value ) else ''
		self.dataObj['영업담당'] = value if (value:=ws['J7'].value ) else '' 


		for rowNo in range(self.tail_start, ws.max_row+1 ):
			value = ws.cell(row=rowNo, column=1).value
			try:
				if ( keyName := Utils.check_str_contains_listElements(value.replace(' ',''), self.checkList) ):
					self.dataObj[keyName] = value if ( value:= ws.cell(row=rowNo, column= 5 ).value) else ''
			except:
				pass

	def get_image(self, ws:Worksheet):
		""" image : pillow_image https://pypi.org/project/openpyxl-image-loader/"""
		#https://stackoverflow.com/questions/43748023/how-to-find-cells-that-contain-images-with-openpyxl
		# Put your sheet in the loader
		image_loader = SheetImageLoader(ws)

		# And get image from specified cell
		try:
			image = image_loader.get(f'A{self.tail_start-1}')
			self.dataObj['pilImage'] = image
		except:
			pass

		# Image now is a Pillow image, so you can do the following
		# image.show()



# class Upload_Excel_생지( Upload_Excel_작지 ):
# 	def __init__(self, **kwargs):
# 		self.init_attributes(**kwargs)
# 		self.process_table_header=  {
# 			2: '적용부품',
# 			3: '적용패널',
# 			5: 'Material',
# 			6: 'Process',
# 			10: '비고',
# 		}
# 		self.dataObj = {'id': -1}
# 		self.process_result = []
# 		self.checkList =  ['고객요청사항','고객성향','특이사항','집중점검항목','검사요청사항']

# 	def run(self) -> dict : #tuple[list]:
# 		clipboard = QP


# 		# self.read(fName)

# 		# ws = self.wb.worksheets[0]

# 		# self.get_table_process_datas(ws)
# 		# self.get_datas(ws)

# 		# self.dataObj['process_fks']= self.process_result
# 		# self.get_image(ws)

# 		# return self.dataObj